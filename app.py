"""
Dell iDRAC Hardware Inventory Dashboard
========================================
Flask backend that proxies Redfish API calls to Dell iDRAC controllers
and returns normalized hardware inventory data.

Architecture:
  Browser  -->  Flask (this server)  -->  iDRAC Redfish API (HTTPS)
  - The browser never talks directly to iDRAC
  - Flask handles auth, SSL, and data normalization
  - Credentials are held only in server memory for the duration of the request

Redfish endpoints queried:
  /redfish/v1/Systems/System.Embedded.1
  /redfish/v1/Systems/System.Embedded.1/Processors
  /redfish/v1/Systems/System.Embedded.1/Memory
  /redfish/v1/Chassis/System.Embedded.1
  /redfish/v1/Chassis/System.Embedded.1/Power
  /redfish/v1/Chassis/System.Embedded.1/Thermal
  /redfish/v1/Chassis/System.Embedded.1/NetworkAdapters
  /redfish/v1/Systems/System.Embedded.1/Storage
  /redfish/v1/UpdateService/FirmwareInventory
"""

import json
import csv
import io
import re
import traceback
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, render_template, Response, send_file
import requests
import urllib3
from openpyxl import Workbook, load_workbook

# Suppress InsecureRequestWarning for self-signed iDRAC certs
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _rf_get(base_url, path, auth, timeout=30):
    """Make an authenticated GET to a Redfish endpoint. Returns parsed JSON or None."""
    url = f"{base_url}{path}"
    try:
        resp = requests.get(url, auth=auth, verify=False, timeout=timeout)
        if resp.status_code == 200:
            return resp.json()
        return None
    except Exception:
        return None


def _collect_members(base_url, collection_path, auth):
    """Fetch a Redfish collection and expand each member."""
    coll = _rf_get(base_url, collection_path, auth)
    if not coll or "Members" not in coll:
        return []

    members = []
    member_urls = [m.get("@odata.id", "") for m in coll["Members"]]

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_rf_get, base_url, url, auth): url for url in member_urls}
        for fut in as_completed(futures):
            data = fut.result()
            if data:
                members.append(data)
    return members


def _safe(d, *keys, default="N/A"):
    """Safely traverse nested dicts."""
    cur = d
    for k in keys:
        if isinstance(cur, dict):
            cur = cur.get(k, default)
        else:
            return default
    return cur if cur not in (None, "", 0) else default


# ---------------------------------------------------------------------------
# Inventory parsers  – each returns a list of normalised row dicts
# ---------------------------------------------------------------------------

def _parse_system(data):
    """Parse /redfish/v1/Systems/System.Embedded.1"""
    if not data:
        return []
    return [{
        "category":    "System",
        "type":        "Server",
        "name":        _safe(data, "Model"),
        "slot":        "N/A",
        "quantity":    1,
        "serial":      _safe(data, "SerialNumber"),
        "part_number": _safe(data, "PartNumber"),
        "firmware":    _safe(data, "BiosVersion"),
        "status":      _safe(data, "Status", "Health"),
        "extra": {
            "Manufacturer": _safe(data, "Manufacturer"),
            "ServiceTag":   _safe(data, "SKU"),
            "HostName":     _safe(data, "HostName"),
            "PowerState":   _safe(data, "PowerState"),
            "MemoryGiB":    _safe(data, "MemorySummary", "TotalSystemMemoryGiB"),
            "ProcessorCount": _safe(data, "ProcessorSummary", "Count"),
            "ProcessorModel": _safe(data, "ProcessorSummary", "Model"),
        }
    }]


def _is_gpu_processor(p):
    """Detect whether a Processors member is a GPU/accelerator rather than a CPU."""
    # Dell iDRAC sets ProcessorType to "GPU" or "OEM" for accelerators
    ptype = str(_safe(p, "ProcessorType", default="")).lower()
    if ptype in ("gpu", "accelerator"):
        return True

    # Check the model/name/description for GPU keywords
    gpu_keywords = [
        "gpu", "nvidia", "amd", "accelerator", "tesla", "a100", "h100",
        "v100", "a30", "a40", "l40", "l4", "radeon", "instinct", "geforce",
        "a2", "a10", "a16", "t4", "display", "graphics",
    ]
    model = str(_safe(p, "Model", default="")).lower()
    name  = str(_safe(p, "Name", default="")).lower()
    desc  = str(_safe(p, "Description", default="")).lower()
    mfr   = str(_safe(p, "Manufacturer", default="")).lower()
    pid   = str(_safe(p, "Id", default="")).lower()
    combined = f"{model} {name} {desc} {mfr} {pid}"
    if any(kw in combined for kw in gpu_keywords):
        return True

    # Dell iDRAC often uses "Video" in the Id for GPUs (e.g., "Video.Slot.3")
    if "video" in pid:
        return True

    # OEM ProcessorType with InstructionSet that is NOT x86/ARM → likely GPU
    iset = str(_safe(p, "InstructionSet", default="")).lower()
    if ptype == "oem" and iset not in ("x86", "x86-64", "ia-64", "arm-a64", "arm-a32"):
        return True

    return False


def _get_processor_slot(p):
    """
    Extract the best slot/location identifier for a processor.
    Priority: Dell OEM Slot → Location → Socket → Id
    Dell iDRAC typically uses Id like 'CPU.Socket.1' or 'Video.Slot.3'.
    """
    # Try Dell OEM location data
    oem_dell = p.get("Oem", {}).get("Dell", {})
    if isinstance(oem_dell, dict):
        # DellProcessor.Connector or DellProcessor.SlotNumber
        dell_proc = oem_dell.get("DellProcessor", {})
        if isinstance(dell_proc, dict):
            connector = dell_proc.get("Connector")
            if connector:
                return str(connector)

    # Try standard Location field (Redfish 2020+)
    location = p.get("Location", {})
    if isinstance(location, dict):
        part_loc = location.get("PartLocation", {})
        if isinstance(part_loc, dict):
            svc_label = part_loc.get("ServiceLabel")
            if svc_label:
                return str(svc_label)
            loc_ordinal = part_loc.get("LocationOrdinalValue")
            loc_type = part_loc.get("LocationType", "")
            if loc_ordinal is not None:
                return f"{loc_type} {loc_ordinal}".strip()

    # Try Socket field
    socket = _safe(p, "Socket")
    if socket != "N/A":
        return socket

    # Fallback to Id (e.g., "CPU.Socket.1", "Video.Slot.3")
    return _safe(p, "Id")


def _parse_processors(members):
    """Parse /Processors members – split CPUs and GPUs/Accelerators."""
    rows = []
    for p in members:
        is_gpu = _is_gpu_processor(p)
        slot = _get_processor_slot(p)

        if is_gpu:
            rows.append({
                "category":    "Accelerator",
                "type":        "GPU",
                "name":        _safe(p, "Model"),
                "slot":        slot,
                "quantity":    1,
                "serial":      _safe(p, "SerialNumber"),
                "part_number": _safe(p, "PartNumber"),
                "firmware":    _safe(p, "FirmwareVersion", default=_safe(p, "ProcessorId", "MicrocodeInfo")),
                "status":      _safe(p, "Status", "Health"),
                "extra": {
                    "Manufacturer":   _safe(p, "Manufacturer"),
                    "ProcessorType":  _safe(p, "ProcessorType"),
                    "MaxSpeedMHz":    _safe(p, "MaxSpeedMHz"),
                    "TotalCores":     _safe(p, "TotalCores"),
                    "TotalThreads":   _safe(p, "TotalThreads"),
                    "InstructionSet": _safe(p, "InstructionSet"),
                }
            })
        else:
            rows.append({
                "category":    "Processor",
                "type":        "CPU",
                "name":        _safe(p, "Model"),
                "slot":        slot,
                "quantity":    _safe(p, "TotalCores", default=1),
                "serial":      _safe(p, "SerialNumber"),
                "part_number": _safe(p, "PartNumber"),
                "firmware":    _safe(p, "ProcessorId", "MicrocodeInfo"),
                "status":      _safe(p, "Status", "Health"),
                "extra": {
                    "Manufacturer":   _safe(p, "Manufacturer"),
                    "MaxSpeedMHz":    _safe(p, "MaxSpeedMHz"),
                    "TotalCores":     _safe(p, "TotalCores"),
                    "TotalThreads":   _safe(p, "TotalThreads"),
                    "InstructionSet": _safe(p, "InstructionSet"),
                }
            })
    return rows


def _parse_memory(members):
    rows = []
    for m in members:
        size = _safe(m, "CapacityMiB")
        size_str = f"{int(size)//1024} GB" if isinstance(size, (int, float)) and size > 0 else _safe(m, "CapacityMiB")
        rows.append({
            "category":    "Memory",
            "type":        _safe(m, "MemoryDeviceType", default="DIMM"),
            "name":        f"{_safe(m, 'Manufacturer')} {size_str} {_safe(m, 'OperatingSpeedMhz', default='')}MHz".strip(),
            "slot":        _safe(m, "DeviceLocator", default=_safe(m, "Id")),
            "quantity":    1,
            "serial":      _safe(m, "SerialNumber"),
            "part_number": _safe(m, "PartNumber"),
            "firmware":    "N/A",
            "status":      _safe(m, "Status", "Health"),
            "extra": {
                "Manufacturer":       _safe(m, "Manufacturer"),
                "CapacityMiB":        size,
                "OperatingSpeedMhz":  _safe(m, "OperatingSpeedMhz"),
                "MemoryType":         _safe(m, "MemoryDeviceType"),
                "RankCount":          _safe(m, "RankCount"),
                "DataWidthBits":      _safe(m, "DataWidthBits"),
                "ErrorCorrection":    _safe(m, "ErrorCorrection"),
            }
        })
    return rows


def _extract_perc_part_number(ctrl, sc):
    """
    Extract part number for Dell PERC/HBA controllers.
    Dell iDRAC stores part numbers in multiple locations:
      1. StorageControllers[].PartNumber  (standard Redfish – often empty on PERC)
      2. Oem.Dell.DellController.PartNumber  (under the StorageController sub-object)
      3. Top-level controller Oem.Dell.DellController fields
      4. Oem.Dell.DellControllerBattery or similar nested objects
    """
    # 1. Standard field on the StorageController sub-object
    pn = sc.get("PartNumber")
    if pn and str(pn).strip() and str(pn).strip() != "N/A":
        return str(pn).strip()

    # 2. Dell OEM under the StorageController sub-object
    sc_oem = sc.get("Oem", {}).get("Dell", {})
    if isinstance(sc_oem, dict):
        for key in ("DellController", "DellStorageController"):
            obj = sc_oem.get(key, {})
            if isinstance(obj, dict):
                pn = obj.get("PartNumber")
                if pn and str(pn).strip():
                    return str(pn).strip()

    # 3. Dell OEM at the top-level controller object
    ctrl_oem = ctrl.get("Oem", {}).get("Dell", {})
    if isinstance(ctrl_oem, dict):
        for key in ("DellController", "DellStorageController", "DellRaidController"):
            obj = ctrl_oem.get(key, {})
            if isinstance(obj, dict):
                pn = obj.get("PartNumber")
                if pn and str(pn).strip():
                    return str(pn).strip()
        # Some iDRAC versions put it directly under Oem.Dell
        pn = ctrl_oem.get("PartNumber")
        if pn and str(pn).strip():
            return str(pn).strip()

    # 4. Top-level controller PartNumber (sometimes populated)
    pn = ctrl.get("PartNumber")
    if pn and str(pn).strip():
        return str(pn).strip()

    return "N/A"


def _parse_storage(base_url, auth):
    """Fetch storage controllers and their drives."""
    rows = []
    controllers = _collect_members(base_url, "/redfish/v1/Systems/System.Embedded.1/Storage", auth)

    for ctrl in controllers:
        ctrl_name = _safe(ctrl, "Name", default=_safe(ctrl, "Id"))
        ctrl_id = _safe(ctrl, "Id", default="")

        # Parse controller itself
        if ctrl.get("StorageControllers"):
            for sc in ctrl["StorageControllers"]:
                part_number = _extract_perc_part_number(ctrl, sc)
                rows.append({
                    "category":    "Storage Controller",
                    "type":        "RAID/HBA",
                    "name":        _safe(sc, "Model", default=_safe(sc, "Name")),
                    "slot":        ctrl_name,
                    "quantity":    1,
                    "serial":      _safe(sc, "SerialNumber"),
                    "part_number": part_number,
                    "firmware":    _safe(sc, "FirmwareVersion"),
                    "status":      _safe(sc, "Status", "Health"),
                    "extra": {
                        "Manufacturer":       _safe(sc, "Manufacturer"),
                        "SupportedDeviceProtocols": _safe(sc, "SupportedDeviceProtocols"),
                        "ControllerId":       ctrl_id,
                    }
                })

        # Parse drives
        drives_list = ctrl.get("Drives", [])
        drive_urls = [d.get("@odata.id", "") for d in drives_list]

        with ThreadPoolExecutor(max_workers=8) as pool:
            futs = {pool.submit(_rf_get, base_url, url, auth): url for url in drive_urls}
            for fut in as_completed(futs):
                drv = fut.result()
                if not drv:
                    continue
                cap = _safe(drv, "CapacityBytes")
                if isinstance(cap, (int, float)) and cap > 0:
                    cap_str = f"{cap / (1024**3):.0f} GB"
                else:
                    cap_str = "N/A"
                rows.append({
                    "category":    "Storage Drive",
                    "type":        _safe(drv, "MediaType", default="Disk"),
                    "name":        f"{_safe(drv, 'Manufacturer', default='')} {_safe(drv, 'Model', default='')} {cap_str}".strip(),
                    "slot":        _safe(drv, "PhysicalLocation", "PartLocation", "ServiceLabel",
                                         default=_safe(drv, "Id")),
                    "quantity":    1,
                    "serial":      _safe(drv, "SerialNumber"),
                    "part_number": _safe(drv, "PartNumber"),
                    "firmware":    _safe(drv, "Revision", default=_safe(drv, "FirmwareVersion")),
                    "status":      _safe(drv, "Status", "Health"),
                    "extra": {
                        "Manufacturer": _safe(drv, "Manufacturer"),
                        "MediaType":    _safe(drv, "MediaType"),
                        "Protocol":     _safe(drv, "Protocol"),
                        "CapacityGB":   cap_str,
                        "RotationSpeedRPM": _safe(drv, "RotationSpeedRPM"),
                        "BlockSizeBytes":   _safe(drv, "BlockSizeBytes"),
                    }
                })
    return rows


def _parse_network_adapters(base_url, auth):
    """Fetch NICs from Chassis NetworkAdapters collection."""
    rows = []

    # Try Chassis-level NetworkAdapters (Dell iDRAC standard path)
    adapters = _collect_members(base_url, "/redfish/v1/Chassis/System.Embedded.1/NetworkAdapters", auth)
    if not adapters:
        # Fallback: try Systems-level EthernetInterfaces
        eths = _collect_members(base_url, "/redfish/v1/Systems/System.Embedded.1/EthernetInterfaces", auth)
        for eth in eths:
            rows.append({
                "category":    "Network",
                "type":        "Ethernet",
                "name":        _safe(eth, "Name", default=_safe(eth, "Id")),
                "slot":        _safe(eth, "Id"),
                "quantity":    1,
                "serial":      _safe(eth, "SerialNumber"),
                "part_number": _safe(eth, "PartNumber"),
                "firmware":    "N/A",
                "status":      _safe(eth, "Status", "Health"),
                "extra": {
                    "MACAddress": _safe(eth, "MACAddress"),
                    "SpeedMbps":  _safe(eth, "SpeedMbps"),
                    "LinkStatus": _safe(eth, "LinkStatus"),
                }
            })
        return rows

    for adapter in adapters:
        adapter_name = _safe(adapter, "Model", default=_safe(adapter, "Name"))
        adapter_id = _safe(adapter, "Id")
        port_count = 0

        # Try to get NetworkPorts or Ports
        ports_link = adapter.get("NetworkPorts", {}).get("@odata.id") or adapter.get("Ports", {}).get("@odata.id")
        if ports_link:
            ports = _collect_members(base_url, ports_link, auth)
            port_count = len(ports)
            for port in ports:
                rows.append({
                    "category":    "Network Port",
                    "type":        _safe(port, "ActiveLinkTechnology", default="Ethernet"),
                    "name":        f"{adapter_name} Port {_safe(port, 'Id')}",
                    "slot":        f"{adapter_id} / {_safe(port, 'Id')}",
                    "quantity":    1,
                    "serial":      _safe(adapter, "SerialNumber"),
                    "part_number": _safe(adapter, "PartNumber"),
                    "firmware":    "N/A",
                    "status":      _safe(port, "Status", "Health"),
                    "extra": {
                        "LinkStatus":     _safe(port, "LinkStatus"),
                        "CurrentSpeedGbps": _safe(port, "CurrentSpeedGbps"),
                        "MACAddress":     _safe(port, "AssociatedNetworkAddresses"),
                    }
                })

        # Adapter-level row
        rows.append({
            "category":    "Network Adapter",
            "type":        "NIC",
            "name":        adapter_name,
            "slot":        adapter_id,
            "quantity":    port_count if port_count else 1,
            "serial":      _safe(adapter, "SerialNumber"),
            "part_number": _safe(adapter, "PartNumber"),
            "firmware":    "N/A",
            "status":      _safe(adapter, "Status", "Health"),
            "extra": {
                "Manufacturer": _safe(adapter, "Manufacturer"),
                "PortCount":    port_count,
            }
        })
    return rows


def _parse_power(chassis_data, base_url, auth):
    """Fetch power supplies from Chassis Power endpoint."""
    rows = []
    power = _rf_get(base_url, "/redfish/v1/Chassis/System.Embedded.1/Power", auth)
    if not power:
        return rows

    for ps in power.get("PowerSupplies", []):
        cap = _safe(ps, "PowerCapacityWatts")
        cap_str = f"{cap}W" if cap != "N/A" else ""
        rows.append({
            "category":    "Power Supply",
            "type":        _safe(ps, "PowerSupplyType", default="PSU"),
            "name":        f"{_safe(ps, 'Manufacturer', default='')} {_safe(ps, 'Model', default='')} {cap_str}".strip(),
            "slot":        _safe(ps, "Name", default=_safe(ps, "MemberId")),
            "quantity":    1,
            "serial":      _safe(ps, "SerialNumber"),
            "part_number": _safe(ps, "PartNumber"),
            "firmware":    _safe(ps, "FirmwareVersion"),
            "status":      _safe(ps, "Status", "Health"),
            "extra": {
                "Manufacturer":      _safe(ps, "Manufacturer"),
                "PowerCapacityWatts": cap,
                "LineInputVoltage":  _safe(ps, "LineInputVoltage"),
                "InputRanges":       _safe(ps, "InputRanges"),
            }
        })
    return rows


def _parse_thermal(base_url, auth, system_model=""):
    """Fetch fans from Chassis Thermal endpoint."""
    rows = []
    
    # Try multiple thermal endpoint paths
    thermal_paths = [
        "/redfish/v1/Chassis/System.Embedded.1/Thermal",
        "/redfish/v1/Chassis/Chassis.Embedded.1/Thermal",
        "/redfish/v1/Systems/System.Embedded.1/Thermal"
    ]
    
    thermal = None
    thermal_path_used = None
    
    for path in thermal_paths:
        try:
            thermal = _rf_get(base_url, path, auth, timeout=15)
            if thermal:
                thermal_path_used = path
                break
        except Exception:
            continue
    
    if not thermal:
        # Try to get available chassis to see what's available
        try:
            chassis_data = _rf_get(base_url, "/redfish/v1/Chassis", auth, timeout=10)
            if chassis_data and "Members" in chassis_data:
                for chassis in chassis_data["Members"]:
                    chassis_url = chassis["@odata.id"]
                    chassis_info = _rf_get(base_url, chassis_url, auth, timeout=10)
                    if chassis_info and "Thermal" in chassis_info:
                        thermal_url = chassis_info["Thermal"]["@odata.id"]
                        thermal = _rf_get(base_url, thermal_url, auth, timeout=15)
                        if thermal:
                            thermal_path_used = thermal_url
                            break
        except Exception:
            pass
        
        if not thermal:
            return rows

    # Determine fan tier based on system model
    fan_tier = "Silver"  # Default tier
    if system_model:
        model_upper = system_model.upper()
        if any(x in model_upper for x in ["R750", "R760", "R7515", "R7615", "R840", "R850", "R960", "R960XA"]):
            fan_tier = "Platinum"
        elif any(x in model_upper for x in ["R640", "R650", "R6515", "R6525", "R740", "R7415", "R7425"]):
            fan_tier = "Gold"
        else:
            fan_tier = "Silver"

    fans = thermal.get("Fans", [])
    
    # If no fans found in the primary Fans array, try alternative structures
    if not fans:
        # Some iDRAC versions use different structures
        if "Temperatures" in thermal:
            # Look for fan data in temperature readings (some systems mix them)
            temp_readings = thermal.get("Temperatures", [])
            for reading in temp_readings:
                if "fan" in str(reading.get("Name", "")).lower() or "fan" in str(reading.get("PhysicalContext", "")).lower():
                    fans.append(reading)
    
    for fan in fans:
        # Extract fan status with better fallbacks
        fan_status = "Unknown"
        if "Status" in fan:
            status_obj = fan["Status"]
            if isinstance(status_obj, dict):
                fan_status = _safe(status_obj, "Health", default=_safe(status_obj, "State", default="Unknown"))
            else:
                fan_status = str(status_obj)
        elif "State" in fan:
            fan_status = _safe(fan, "State")
        elif "Health" in fan:
            fan_status = _safe(fan, "Health")
        
        # Extract fan speed/RPM
        reading_rpm = _safe(fan, "Reading")
        if reading_rpm == "N/A" or reading_rpm == "":
            # Try alternative RPM fields
            reading_rpm = _safe(fan, "SpeedRPM", default=_safe(fan, "CurrentSpeed", default=_safe(fan, "Speed")))
        
        # Extract fan name with better fallbacks
        fan_name = _safe(fan, "Name", default=_safe(fan, "FanName", default=_safe(fan, "MemberId")))
        if not fan_name or fan_name == "N/A":
            fan_name = f"Fan {_safe(fan, 'MemberId', default='Unknown')}"
        
        rows.append({
            "category":    "Fan",
            "type":        "Cooling Fan",
            "name":        fan_name,
            "slot":        _safe(fan, "MemberId", default=_safe(fan, "PhysicalContext", default=fan_name)),
            "quantity":    1,
            "serial":      _safe(fan, "SerialNumber"),
            "part_number": _safe(fan, "PartNumber"),
            "firmware":    "N/A",
            "status":      fan_status,
            "extra": {
                "ReadingRPM":     reading_rpm,
                "ReadingUnits":   _safe(fan, "ReadingUnits", default="RPM"),
                "PhysicalContext": _safe(fan, "PhysicalContext"),
                "Description":    _safe(fan, "Description"),
                "FanTier":        fan_tier,
                "ThermalPath":    thermal_path_used,
                "HotPluggable":   _safe(fan, "HotPluggable"),
                "Redundancy":     _safe(fan, "Redundancy"),
            }
        })
    
    # Add debug info if no fans found
    if not rows:
        # Add a debug entry to help troubleshoot
        rows.append({
            "category":    "Fan",
            "type":        "Debug Info",
            "name":        "No Fan Data Found",
            "slot":        f"Endpoint: {thermal_path_used or 'None'}",
            "quantity":    0,
            "serial":      "N/A",
            "part_number": "N/A",
            "firmware":    "N/A",
            "status":      "Warning",
            "extra": {
                "AvailableKeys": list(thermal.keys()) if isinstance(thermal, dict) else "Non-dict response",
                "ThermalPath": thermal_path_used,
                "FanTier": fan_tier,
            }
        })
    
    return rows


def _parse_racadm_hwinventory(host, username, password):
    """Fetch hardware inventory via racadm hwinventory command."""
    import subprocess
    rows = []

    try:
        # Run racadm hwinventory command
        # Note: This requires racadm to be installed and configured on the system
        # Or SSH access to iDRAC to run racadm commands
        cmd = [
            "racadm", "-r", host,
            "-u", username,
            "-p", password,
            "hwinventory"
        ]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)

        if result.returncode != 0:
            return rows

        # Parse racadm hwinventory output
        # The output format varies by iDRAC version, but typically has sections
        current_section = None
        for line in result.stdout.split('\n'):
            line = line.strip()

            # Detect section headers
            if line.startswith('[') and line.endswith(']'):
                current_section = line[1:-1]
                continue

            # Parse key-value pairs
            if '=' in line and current_section:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()

                # Normalize component type based on section
                component_type = _normalize_component_type(current_section)

                rows.append({
                    "component_type": component_type,
                    "device_description": key,
                    "part_number": value,
                    "serial_number": "",
                    "source_interface": "racadm"
                })

    except subprocess.TimeoutExpired:
        return rows
    except FileNotFoundError:
        # racadm not found - this is expected if not installed locally
        return rows
    except Exception as e:
        return rows

    return rows


def _parse_ipmi_fru(host, username, password):
    """Fetch FRU data via IPMI."""
    import subprocess
    rows = []

    try:
        # Run ipmitool fru command
        # Note: This requires ipmitool to be installed
        # And IPMI access to be configured on the server
        cmd = [
            "ipmitool", "-H", host,
            "-U", username,
            "-P", password,
            "fru"
        ]

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)

        if result.returncode != 0:
            return rows

        # Parse IPMI FRU output
        current_device = None
        fru_data = {}

        for line in result.stdout.split('\n'):
            line = line.strip()

            # Detect device headers
            if line.startswith('FRU Device Description'):
                # Save previous device if exists
                if current_device and fru_data:
                    rows.append({
                        "component_type": _normalize_component_type(current_device),
                        "device_description": fru_data.get('Product Name', current_device),
                        "part_number": fru_data.get('Board Part Number', fru_data.get('Product Part Number', '')),
                        "serial_number": fru_data.get('Board Serial', fru_data.get('Product Serial', '')),
                        "source_interface": "ipmi"
                    })
                # Start new device
                current_device = line.split(':', 1)[1].strip() if ':' in line else line
                fru_data = {}
                continue

            # Parse FRU fields
            if ':' in line and current_device:
                key, value = line.split(':', 1)
                key = key.strip()
                value = value.strip()
                fru_data[key] = value

        # Don't forget the last device
        if current_device and fru_data:
            rows.append({
                "component_type": _normalize_component_type(current_device),
                "device_description": fru_data.get('Product Name', current_device),
                "part_number": fru_data.get('Board Part Number', fru_data.get('Product Part Number', '')),
                "serial_number": fru_data.get('Board Serial', fru_data.get('Product Serial', '')),
                "source_interface": "ipmi"
            })

    except subprocess.TimeoutExpired:
        return rows
    except FileNotFoundError:
        # ipmitool not found - this is expected if not installed
        return rows
    except Exception as e:
        return rows

    return rows


def _normalize_component_type(raw_type):
    """Normalize component type across interfaces."""
    if not raw_type:
        return "Unknown"

    type_map = {
        # CPU variations
        "CPU": "Processor",
        "Processor": "Processor",
        "PROC": "Processor",

        # Memory variations
        "DIMM": "Memory",
        "Memory": "Memory",
        "RAM": "Memory",
        "DRAM": "Memory",

        # Network variations
        "NIC": "Network Adapter",
        "Network Adapter": "Network Adapter",
        "Ethernet": "Network Adapter",
        "Network": "Network Adapter",

        # Storage variations
        "Disk": "Storage Drive",
        "Storage Drive": "Storage Drive",
        "HDD": "Storage Drive",
        "SSD": "Storage Drive",
        "NVMe": "Storage Drive",

        # Power variations
        "PSU": "Power Supply",
        "Power Supply": "Power Supply",
        "Power": "Power Supply",

        # Fan variations
        "Fan": "Fan",
        "Cooling Fan": "Fan",

        # System board variations
        "System Board": "System Board",
        "Motherboard": "System Board",
        "Main Board": "System Board",
    }

    # Case-insensitive lookup
    upper_type = raw_type.upper()
    for key, value in type_map.items():
        if key.upper() in upper_type or upper_type in key.upper():
            return value

    # Return original if no match
    return raw_type


def _normalize_inventory_data(raw_data, source):
    """Normalize inventory data to standard format."""
    normalized = []

    for item in raw_data:
        normalized.append({
            "component_type": _normalize_component_type(item.get("category", item.get("component_type", ""))),
            "device_description": item.get("name", item.get("device_description", "")),
            "part_number": item.get("part_number", ""),
            "serial_number": item.get("serial", item.get("serial_number", "")),
            "source_interface": source
        })

    return normalized


def _compare_inventory_across_interfaces(redfish_data, racadm_data, ipmi_data):
    """Compare inventory data across all three interfaces."""
    comparison_results = {}

    # Create a map of component_type -> part_number -> list of sources
    component_map = {}

    # Add Redfish data
    for item in redfish_data:
        key = (item["component_type"], item["part_number"])
        if key not in component_map:
            component_map[key] = {"redfish": None, "racadm": None, "ipmi": None}
        component_map[key]["redfish"] = item

    # Add racadm data
    for item in racadm_data:
        key = (item["component_type"], item["part_number"])
        if key not in component_map:
            component_map[key] = {"redfish": None, "racadm": None, "ipmi": None}
        component_map[key]["racadm"] = item

    # Add IPMI data
    for item in ipmi_data:
        key = (item["component_type"], item["part_number"])
        if key not in component_map:
            component_map[key] = {"redfish": None, "racadm": None, "ipmi": None}
        component_map[key]["ipmi"] = item

    # Generate comparison results
    for key, sources in component_map.items():
        component_type, part_number = key

        # Determine status
        has_redfish = sources["redfish"] is not None
        has_racadm = sources["racadm"] is not None
        has_ipmi = sources["ipmi"] is not None

        if has_redfish and has_racadm and has_ipmi:
            status = "match"
            # Check for mismatches in descriptions
            if (sources["redfish"]["device_description"] != sources["racadm"].get("device_description") or
                sources["redfish"]["device_description"] != sources["ipmi"].get("device_description")):
                status = "description_mismatch"
        elif has_redfish and has_racadm:
            status = "missing_ipmi"
        elif has_redfish and has_ipmi:
            status = "missing_racadm"
        elif has_racadm and has_ipmi:
            status = "missing_redfish"
        elif has_redfish:
            status = "redfish_only"
        elif has_racadm:
            status = "racadm_only"
        elif has_ipmi:
            status = "ipmi_only"
        else:
            status = "unknown"

        comparison_results[f"{component_type}_{part_number}"] = {
            "component_type": component_type,
            "part_number": part_number,
            "redfish": {
                "device_description": sources["redfish"]["device_description"] if sources["redfish"] else None,
                "serial_number": sources["redfish"]["serial_number"] if sources["redfish"] else None
            } if sources["redfish"] else None,
            "racadm": {
                "device_description": sources["racadm"]["device_description"] if sources["racadm"] else None,
                "serial_number": sources["racadm"]["serial_number"] if sources["racadm"] else None
            } if sources["racadm"] else None,
            "ipmi": {
                "device_description": sources["ipmi"]["device_description"] if sources["ipmi"] else None,
                "serial_number": sources["ipmi"]["serial_number"] if sources["ipmi"] else None
            } if sources["ipmi"] else None,
            "status": status
        }

    return comparison_results


def _parse_gpu_pcie(base_url, auth, known_gpu_serials=None):
    """
    Fetch PCIe devices – filter for GPUs/accelerators + list others.
    Skips GPUs whose serial numbers are already in known_gpu_serials
    (to avoid duplicating entries already captured from /Processors).
    """
    rows = []
    known = set(known_gpu_serials or [])
    known.discard("N/A")
    known.discard("")

    # Try the PCIeDevices collection under Systems
    devices = _collect_members(base_url, "/redfish/v1/Systems/System.Embedded.1/PCIeDevices", auth)

    # Fallback: try Chassis PCIeDevices
    if not devices:
        chassis = _rf_get(base_url, "/redfish/v1/Chassis/System.Embedded.1", auth)
        if chassis:
            pcie_links = chassis.get("PCIeDevices", [])
            urls = [p.get("@odata.id", "") for p in pcie_links]
            with ThreadPoolExecutor(max_workers=8) as pool:
                futs = {pool.submit(_rf_get, base_url, u, auth): u for u in urls}
                for fut in as_completed(futs):
                    d = fut.result()
                    if d:
                        devices.append(d)

    gpu_keywords = ["gpu", "nvidia", "amd", "accelerator", "tesla", "a100", "h100",
                    "v100", "a30", "a40", "l40", "l4", "radeon", "instinct", "geforce",
                    "a2", "a10", "a16", "t4", "display", "graphics"]

    for dev in devices:
        name = _safe(dev, "Model", default=_safe(dev, "Name", default=_safe(dev, "Id")))
        desc = _safe(dev, "Description", default="")
        device_class = _safe(dev, "DeviceClass", default="")
        combined = f"{name} {desc} {device_class}".lower()

        is_gpu = any(kw in combined for kw in gpu_keywords)

        # Skip GPUs already captured from /Processors
        if is_gpu:
            serial = _safe(dev, "SerialNumber")
            dev_name = name.lower()
            if serial in known:
                continue
            # Also check by name overlap with known GPU names
            cat = "Accelerator"
        else:
            cat = "PCIe Device"

        # Extract slot from PCIeFunctions or Location
        slot = _safe(dev, "Id")
        oem = dev.get("Oem", {}).get("Dell", {})
        if isinstance(oem, dict):
            for key in ("DellPCIeDevice", "DellPCIeFunction"):
                obj = oem.get(key, {})
                if isinstance(obj, dict):
                    sl = obj.get("SlotNumber") or obj.get("Slot")
                    if sl is not None:
                        slot = f"Slot {sl}"
                        break

        rows.append({
            "category":    cat,
            "type":        "GPU" if is_gpu else _safe(dev, "DeviceClass", default="PCIe"),
            "name":        name,
            "slot":        slot,
            "quantity":    1,
            "serial":      _safe(dev, "SerialNumber"),
            "part_number": _safe(dev, "PartNumber"),
            "firmware":    _safe(dev, "FirmwareVersion"),
            "status":      _safe(dev, "Status", "Health"),
            "extra": {
                "Manufacturer": _safe(dev, "Manufacturer"),
                "DeviceClass":  device_class,
                "Description":  desc,
            }
        })
    return rows


def _parse_firmware(base_url, auth):
    """Fetch firmware inventory from UpdateService."""
    rows = []
    members = _collect_members(base_url, "/redfish/v1/UpdateService/FirmwareInventory", auth)
    for fw in members:
        rows.append({
            "category":    "Firmware",
            "type":        "Firmware",
            "name":        _safe(fw, "Name"),
            "slot":        _safe(fw, "Id"),
            "quantity":    1,
            "serial":      "N/A",
            "part_number": "N/A",
            "firmware":    _safe(fw, "Version"),
            "status":      _safe(fw, "Status", "Health"),
            "extra": {
                "Updateable":    _safe(fw, "Updateable"),
                "SoftwareId":    _safe(fw, "SoftwareId"),
                "ReleaseDate":   _safe(fw, "ReleaseDate"),
            }
        })
    return rows


# ---------------------------------------------------------------------------
# Orchestrator – calls all parsers and merges results
# ---------------------------------------------------------------------------

def fetch_full_inventory(host, username, password):
    """
    Main function: queries all Redfish endpoints in parallel and returns
    the combined inventory rows plus system summary.
    """
    base_url = f"https://{host}"
    auth = (username, password)

    # 1. Validate connectivity and credentials
    system = _rf_get(base_url, "/redfish/v1/Systems/System.Embedded.1", auth, timeout=15)
    if system is None:
        # Try a basic Redfish root to differentiate auth vs. unreachable
        root = _rf_get(base_url, "/redfish/v1/", auth, timeout=10)
        if root is None:
            return None, "unreachable", "Cannot connect to iDRAC at the given IP address. Verify the IP and network connectivity."
        else:
            return None, "auth_failed", "Authentication failed. Please check your username and password."

    # 2. Fetch all component data in parallel (except GPU/PCIe which needs processor results)
    inventory = []
    errors = []

    # Get system model for fan tier determination
    system_model = system.get("Model", "") if system else ""

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {
            pool.submit(_parse_system, system): "System",
            pool.submit(_collect_members, base_url, "/redfish/v1/Systems/System.Embedded.1/Processors", auth): "Processors_raw",
            pool.submit(_collect_members, base_url, "/redfish/v1/Systems/System.Embedded.1/Memory", auth): "Memory_raw",
            pool.submit(_parse_storage, base_url, auth): "Storage",
            pool.submit(_parse_network_adapters, base_url, auth): "Network",
            pool.submit(_rf_get, base_url, "/redfish/v1/Chassis/System.Embedded.1", auth): "Chassis_raw",
            pool.submit(_parse_thermal, base_url, auth, system_model): "Thermal",
            pool.submit(_parse_firmware, base_url, auth): "Firmware",
        }

        results = {}
        for fut in as_completed(futures):
            label = futures[fut]
            try:
                results[label] = fut.result()
            except Exception as e:
                errors.append(f"{label}: {str(e)}")
                results[label] = [] if label != "Chassis_raw" else None

    # Post-process processors (splits CPUs and GPUs/Accelerators)
    proc_rows = _parse_processors(results.get("Processors_raw", []))

    # Collect serial numbers of GPUs found in /Processors to avoid duplicates in PCIe
    known_gpu_serials = [
        r["serial"] for r in proc_rows
        if r["category"] == "Accelerator" and r["serial"] != "N/A"
    ]

    # Now fetch PCIe devices with deduplication
    try:
        gpu_pcie_rows = _parse_gpu_pcie(base_url, auth, known_gpu_serials=known_gpu_serials)
    except Exception as e:
        errors.append(f"GPU_PCIe: {str(e)}")
        gpu_pcie_rows = []

    # Assemble final inventory
    inventory.extend(results.get("System", []))
    inventory.extend(proc_rows)
    inventory.extend(_parse_memory(results.get("Memory_raw", [])))
    inventory.extend(results.get("Storage", []))
    inventory.extend(results.get("Network", []))
    inventory.extend(_parse_power(results.get("Chassis_raw"), base_url, auth))
    inventory.extend(results.get("Thermal", []))
    inventory.extend(gpu_pcie_rows)
    inventory.extend(results.get("Firmware", []))

    # Build summary
    summary = {
        "host": host,
        "model": _safe(system, "Model"),
        "serviceTag": _safe(system, "SKU"),
        "serialNumber": _safe(system, "SerialNumber"),
        "biosVersion": _safe(system, "BiosVersion"),
        "powerState": _safe(system, "PowerState"),
        "totalMemoryGiB": _safe(system, "MemorySummary", "TotalSystemMemoryGiB"),
        "processorModel": _safe(system, "ProcessorSummary", "Model"),
        "processorCount": _safe(system, "ProcessorSummary", "Count"),
        "totalComponents": len(inventory),
        "warnings": errors,
    }

    return {"summary": summary, "inventory": inventory}, "ok", ""


# ---------------------------------------------------------------------------
# Flask Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/inventory", methods=["POST"])
def api_inventory():
    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid request body"}), 400

    host = (body.get("host") or "").strip()
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""

    if not host or not username or not password:
        return jsonify({"error": "Host, username, and password are all required."}), 400

    try:
        result, status, message = fetch_full_inventory(host, username, password)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected error: {str(e)}"}), 500

    if status == "unreachable":
        return jsonify({"error": message, "code": "UNREACHABLE"}), 502
    elif status == "auth_failed":
        return jsonify({"error": message, "code": "AUTH_FAILED"}), 401
    else:
        return jsonify(result), 200


@app.route("/api/export-csv", methods=["POST"])
def export_csv():
    """Accept inventory JSON and return a CSV file."""
    body = request.get_json(silent=True)
    if not body or "inventory" not in body:
        return jsonify({"error": "No inventory data provided"}), 400

    rows = body["inventory"]
    si = io.StringIO()
    fieldnames = ["category", "type", "name", "slot", "quantity", "serial", "part_number", "firmware", "status"]
    writer = csv.DictWriter(si, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    output = si.getvalue()
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=idrac_inventory.csv"}
    )


# ---------------------------------------------------------------------------
# BOM / Excel Comparison
# ---------------------------------------------------------------------------

def _normalize_pn(pn):
    """Normalize a part number for comparison: strip, uppercase, collapse whitespace."""
    if not pn or str(pn).strip().upper() in ("N/A", "NONE", ""):
        return ""
    return re.sub(r'\s+', ' ', str(pn).strip().upper())


def _pn_match_keys(pn_norm):
    """
    Return a set of match keys derived from a normalized part number.

    System inventory may show '0VJWVJ' while the Excel BOM lists 'VJWVJ'.
    To handle this we generate keys by:
      1. The normalised string itself            → '0VJWVJ'
      2. Leading-zero-stripped version            → 'VJWVJ'
      3. All hyphens/dashes removed               → '0VJWVJ' (no-op here)
      4. Leading-zero-stripped + hyphens removed  → 'VJWVJ'
    All keys are uppercase (already normalised).
    """
    if not pn_norm:
        return set()
    keys = {pn_norm}
    stripped = pn_norm.lstrip("0")
    if stripped:
        keys.add(stripped)
    no_dash = pn_norm.replace("-", "")
    keys.add(no_dash)
    stripped_no_dash = no_dash.lstrip("0")
    if stripped_no_dash:
        keys.add(stripped_no_dash)
    return keys


# Map user-friendly type names to inventory categories
_TYPE_ALIAS = {
    "CPU":     ["Processor"],
    "DIMM":    ["Memory"],
    "MEMORY":  ["Memory"],
    "GPU":     ["Accelerator", "GPU / Accelerator", "PCIe Device"],
    "NIC":     ["Network Adapter", "Network Port", "Network"],
    "DISK":    ["Storage Drive"],
    "DRIVE":   ["Storage Drive"],
    "SSD":     ["Storage Drive"],
    "HDD":     ["Storage Drive"],
    "STORAGE": ["Storage Controller", "Storage Drive"],
    "RAID":    ["Storage Controller"],
    "PERC":    ["Storage Controller"],
    "HBA":     ["Storage Controller"],
    "PSU":     ["Power Supply"],
    "FAN":     ["Fan"],
    "SYSTEM":  ["System"],
}


def _resolve_categories(comp_type_raw):
    """Given a user-entered component type, return matching inventory category names."""
    ct = str(comp_type_raw or "").strip().upper()
    if ct in _TYPE_ALIAS:
        return _TYPE_ALIAS[ct]
    # Try partial match
    for alias, cats in _TYPE_ALIAS.items():
        if alias in ct or ct in alias:
            return cats
    # Fallback: try to match directly against category names
    return [comp_type_raw.strip()] if comp_type_raw else []


def _parse_excel(file_stream, filename):
    """
    Parse an uploaded Excel file (.xlsx / .xls) – any filename is accepted.

    Logic:
      1. Look for a worksheet named "lab build sheet" (case-insensitive).
      2. If not found, scan every worksheet for a header row containing
         an "ASSY DPN" or "Part Number" column.  Use the first match.
      3. If still nothing, try the active sheet as a last resort.
      4. Optionally pick up companion columns in the same header row:
         Component Type, Quantity, Slot, Description.
      5. Parse all data rows below the header.

    Returns (rows_list, error_string).
    """
    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
    except Exception as e:
        return None, f"Cannot read Excel file: {str(e)}"

    ASSY_DPN_PATTERNS = ["assy dpn", "assy_dpn", "assydpn", "assy  dpn"]
    PART_NUM_PATTERNS = ["part number", "part_number", "part no", "part#", "pn", "partnumber", "part no."]
    ALL_PN_PATTERNS   = ASSY_DPN_PATTERNS + PART_NUM_PATTERNS
    SCAN_LIMIT = 30

    def _scan_sheet_for_header(sheet):
        """Return (row_idx, row_cells, col_idx) for the first PN header found, or (None,None,None)."""
        for rn, row in enumerate(sheet.iter_rows(min_row=1, max_row=SCAN_LIMIT, values_only=True), start=1):
            if row is None:
                continue
            for ci, cv in enumerate(row):
                if cv is None:
                    continue
                norm = re.sub(r'\s+', ' ', str(cv).strip().lower())
                if norm in ALL_PN_PATTERNS:
                    return rn, row, ci
        return None, None, None

    ws = None
    header_row_idx = None
    header_cells = None
    assy_col_idx = None

    # --- Step 1: prefer a sheet named "lab build sheet" ---
    for name in wb.sheetnames:
        if "lab build sheet" in name.lower():
            ws = wb[name]
            header_row_idx, header_cells, assy_col_idx = _scan_sheet_for_header(ws)
            if assy_col_idx is not None:
                break
            ws = None  # name matched but no header; keep looking
            break

    # --- Step 2: scan every sheet for a PN header ---
    if assy_col_idx is None:
        for name in wb.sheetnames:
            candidate = wb[name]
            h_row, h_cells, h_col = _scan_sheet_for_header(candidate)
            if h_col is not None:
                ws = candidate
                header_row_idx, header_cells, assy_col_idx = h_row, h_cells, h_col
                break

    # --- Step 3: last resort – active sheet ---
    if assy_col_idx is None and wb.active is not None:
        ws = wb.active
        header_row_idx, header_cells, assy_col_idx = _scan_sheet_for_header(ws)

    if ws is None or assy_col_idx is None:
        sheets = ", ".join(wb.sheetnames) if hasattr(wb, 'sheetnames') else "(unknown)"
        wb.close()
        return None, (
            f"Could not find an 'ASSY DPN' or 'Part Number' column in any sheet. "
            f"Scanned the first {SCAN_LIMIT} rows of each. "
            f"Available sheets: {sheets}"
        )

    sheet_used = ws.title

    # --- Step 4: map optional companion columns in the same header row ---
    # The anchor column (assy_col_idx) could be ASSY DPN or Part Number.
    # Figure out which, then look for the other in the remaining columns.
    anchor_norm = re.sub(r'\s+', ' ', str(header_cells[assy_col_idx]).strip().lower())
    if anchor_norm in ASSY_DPN_PATTERNS:
        col_map = {"assy_dpn": assy_col_idx}
        other_pn_aliases = PART_NUM_PATTERNS
        other_pn_key = "alt_part_number"
    else:
        col_map = {"alt_part_number": assy_col_idx}
        other_pn_aliases = ASSY_DPN_PATTERNS
        other_pn_key = "assy_dpn"

    COMPANION_ALIASES = {
        "component_type": ["component type", "component", "type", "category", "comp type", "comp_type"],
        "quantity":       ["quantity", "qty", "count", "expected quantity", "expected qty", "expected_quantity"],
        "slot":           ["slot", "location", "slot / location", "slot/location"],
        "description":    ["description", "desc", "notes", "name", "model"],
    }

    for col_idx, cell_val in enumerate(header_cells):
        if cell_val is None or col_idx == assy_col_idx:
            continue
        norm = re.sub(r'\s+', ' ', str(cell_val).strip().lower())
        # Check for the other PN column
        if other_pn_key not in col_map and norm in other_pn_aliases:
            col_map[other_pn_key] = col_idx
            continue
        for field, aliases in COMPANION_ALIASES.items():
            if norm in aliases and field not in col_map:
                col_map[field] = col_idx
                break

    # --- Step 5: parse data rows below the header ---
    rows = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if row is None:
            continue

        assy_dpn = ""
        if "assy_dpn" in col_map and col_map["assy_dpn"] < len(row):
            raw = row[col_map["assy_dpn"]]
            assy_dpn = str(raw).strip() if raw else ""

        alt_pn = ""
        if "alt_part_number" in col_map and col_map["alt_part_number"] < len(row):
            alt_raw = row[col_map["alt_part_number"]]
            alt_pn = str(alt_raw).strip() if alt_raw else ""

        # Skip rows where neither part number column has a value
        if not assy_dpn and not alt_pn:
            continue

        comp_type = ""
        if "component_type" in col_map and col_map["component_type"] < len(row):
            comp_type = str(row[col_map["component_type"]] or "").strip()

        qty = None
        if "quantity" in col_map and col_map["quantity"] < len(row):
            qty_raw = row[col_map["quantity"]]
            try:
                qty = int(qty_raw) if qty_raw is not None and str(qty_raw).strip() != "" else None
            except (ValueError, TypeError):
                qty = None

        slot_val = ""
        if "slot" in col_map and col_map["slot"] < len(row):
            slot_val = str(row[col_map["slot"]] or "").strip()

        desc_val = ""
        if "description" in col_map and col_map["description"] < len(row):
            desc_val = str(row[col_map["description"]] or "").strip()

        rows.append({
            "row":              row_num,
            "component_type":   comp_type,
            "assy_dpn":         assy_dpn,
            "part_number":      alt_pn,
            "quantity":         qty,
            "slot":             slot_val,
            "description":      desc_val,
        })

    wb.close()

    if not rows:
        return None, (
            f"No data rows with part numbers found below the 'ASSY DPN' header "
            f"(row {header_row_idx}) in sheet '{sheet_used}'."
        )

    return rows, None


def compare_inventory(excel_rows, inventory):
    """
    Compare Excel BOM entries against live inventory.
    Returns a list of comparison result dicts + summary counters.

    Matching strategy:
      Each Excel row may have two part numbers: ASSY DPN and Part Number.
      For every inventory part number we generate multiple "match keys"
      (original, leading-zeros-stripped, dashes-removed, both).
      We do the same for each Excel PN and look for any overlap.
      If either Excel PN matches an inventory row, it's a MATCHED.
    """
    # Build lookup: match_key → list of inventory rows
    pn_index = defaultdict(list)
    for inv_row in inventory:
        npn = _normalize_pn(inv_row.get("part_number"))
        if npn:
            for key in _pn_match_keys(npn):
                pn_index[key].append(inv_row)

    results = []
    summary = {"matched": 0, "not_found": 0, "qty_match": 0, "qty_mismatch": 0, "total": len(excel_rows)}

    for ex in excel_rows:
        ex_assy_raw = ex.get("assy_dpn", "")
        ex_pn_raw   = ex.get("part_number", "")
        ex_assy_norm = _normalize_pn(ex_assy_raw)
        ex_pn_norm   = _normalize_pn(ex_pn_raw)
        ex_type      = ex["component_type"]
        ex_qty       = ex["quantity"]

        if not ex_assy_norm and not ex_pn_norm:
            results.append({
                "excel_row":        ex["row"],
                "component_type":   ex_type,
                "assy_dpn":         ex_assy_raw,
                "part_number":      ex_pn_raw,
                "description":      ex.get("description", ""),
                "detected_parts":   [],
                "detected_qty":     0,
                "expected_qty":     ex_qty,
                "match_status":     "NOT_FOUND",
                "qty_status":       "N/A",
                "detail":           "Both ASSY DPN and Part Number are empty",
            })
            summary["not_found"] += 1
            continue

        # Try all match keys from both Excel PNs
        matched_inv = []
        seen_ids = set()
        for norm_pn in (ex_assy_norm, ex_pn_norm):
            if not norm_pn:
                continue
            for key in _pn_match_keys(norm_pn):
                for row in pn_index.get(key, []):
                    row_id = id(row)
                    if row_id not in seen_ids:
                        seen_ids.add(row_id)
                        matched_inv.append(row)

        # Optionally filter by component type if provided
        target_cats = _resolve_categories(ex_type) if ex_type else []
        if target_cats and matched_inv:
            filtered = [m for m in matched_inv if m.get("category") in target_cats]
            if filtered:
                matched_inv = filtered

        detected_qty = len(matched_inv)
        detected_pns = list({m.get("part_number", "N/A") for m in matched_inv})
        detected_names = list({m.get("name", "N/A") for m in matched_inv})

        # Use inventory Name/Model as component type when matched.
        # Deduplicate: if one name is a prefix/substring of another, keep
        # only the shorter base name (e.g. keep "Broadcom BCM57412 …NIC"
        # and drop "Broadcom BCM57412 …NIC Port NIC.Slot.5-1").
        if matched_inv:
            unique = [n for n in detected_names if n and n != "N/A"]
            # Sort shortest first, then drop any name that contains a shorter one
            unique.sort(key=len)
            pruned = []
            for n in unique:
                nl = n.lower()
                if not any(nl.startswith(p.lower()) or p.lower() in nl for p in pruned):
                    pruned.append(n)
            comp_label = "; ".join(pruned) if pruned else ex_type
        else:
            comp_label = ex_type

        # Match status: only MATCHED or NOT_FOUND
        if not matched_inv:
            status = "NOT_FOUND"
            summary["not_found"] += 1
        else:
            status = "MATCHED"
            summary["matched"] += 1

        # Qty status: separate column
        if not matched_inv:
            qty_status = "N/A"
            detail = "Part number not found in system inventory"
        elif ex_qty is not None and detected_qty != ex_qty:
            qty_status = "QTY_MISMATCH"
            detail = f"Expected {ex_qty}, detected {detected_qty}"
            summary["qty_mismatch"] += 1
        elif ex_qty is not None and detected_qty == ex_qty:
            qty_status = "QTY_MATCH"
            detail = ""
            summary["qty_match"] += 1
        else:
            # No expected qty in Excel – show detected count, mark as match
            qty_status = "QTY_MATCH"
            detail = f"Detected {detected_qty} (no expected qty specified)"
            summary["qty_match"] += 1

        results.append({
            "excel_row":         ex["row"],
            "component_type":    comp_label,
            "assy_dpn":          ex_assy_raw,
            "part_number":       ex_pn_raw,
            "description":       ex.get("description", ""),
            "detected_parts":    detected_pns,
            "detected_names":    detected_names,
            "detected_qty":      detected_qty,
            "expected_qty":      ex_qty,
            "match_status":      status,
            "qty_status":        qty_status,
            "detail":            detail,
        })

    return results, summary


@app.route("/api/compare", methods=["POST"])
def api_compare():
    """
    Accept a multipart form with:
      - file: the Excel BOM file
      - inventory: JSON string of the current inventory array
    Returns comparison results.
    """
    if "file" not in request.files:
        return jsonify({"error": "No Excel file uploaded."}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected."}), 400

    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ("xlsx", "xls"):
        return jsonify({"error": f"Invalid file format '.{ext}'. Please upload an .xlsx or .xls file."}), 400

    inv_json = request.form.get("inventory")
    if not inv_json:
        return jsonify({"error": "No inventory data provided. Fetch inventory first."}), 400

    try:
        inventory = json.loads(inv_json)
    except (json.JSONDecodeError, TypeError):
        return jsonify({"error": "Invalid inventory data format."}), 400

    if not isinstance(inventory, list) or not inventory:
        return jsonify({"error": "Inventory is empty. Fetch inventory before comparing."}), 400

    # Parse Excel in memory
    file_bytes = io.BytesIO(f.read())
    excel_rows, parse_err = _parse_excel(file_bytes, f.filename)
    if parse_err:
        return jsonify({"error": parse_err}), 400

    # Run comparison
    try:
        results, summary = compare_inventory(excel_rows, inventory)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Comparison failed: {str(e)}"}), 500

    return jsonify({"results": results, "summary": summary}), 200


@app.route("/api/export-comparison-csv", methods=["POST"])
def export_comparison_csv():
    """Accept comparison results JSON and return a CSV file."""
    body = request.get_json(silent=True)
    if not body or "results" not in body:
        return jsonify({"error": "No comparison results provided."}), 400

    rows = body["results"]
    si = io.StringIO()
    fieldnames = [
        "component_type", "excel_part_number", "detected_parts",
        "match_status", "qty_status", "detected_qty", "expected_qty", "detail", "description"
    ]
    writer = csv.DictWriter(si, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        r = dict(row)
        # Flatten list fields
        if isinstance(r.get("detected_parts"), list):
            r["detected_parts"] = "; ".join(r["detected_parts"])
        writer.writerow(r)

    return Response(
        si.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bom_comparison_results.csv"}
    )


@app.route("/api/lc-logs", methods=["POST"])
def fetch_lc_logs():
    """Fetch the last 50 LC (Lifecycle Controller) logs from iDRAC."""
    try:
        data = request.get_json()
        host = data.get("host")
        username = data.get("username")
        password = data.get("password")

        if not all([host, username, password]):
            return jsonify({"error": "Missing credentials"}), 400

        auth = (username, password)
        base_url = f"https://{host}"

        # Try multiple LC log endpoint paths for different iDRAC versions
        lc_log_paths = [
            "/redfish/v1/Managers/iDRAC.Embedded.1/Logs/LCLog",
            "/redfish/v1/Managers/iDRAC.Embedded.1/Logs/LcLog",
            "/redfish/v1/Managers/iDRAC.Embedded.1/LogServices/LcLog/Entries",
            "/redfish/v1/Systems/System.Embedded.1/Logs/LcLog",
            "/redfish/v1/Chassis/System.Embedded.1/Logs/LcLog"
        ]
        
        lc_log_data = None
        lc_log_path_used = None
        
        for path in lc_log_paths:
            try:
                lc_log_data = _rf_get(base_url, path, auth, timeout=15)
                if lc_log_data:
                    lc_log_path_used = path
                    break
            except Exception:
                continue

        if not lc_log_data:
            # Try to get available log services to show what's available
            try:
                managers_data = _rf_get(base_url, "/redfish/v1/Managers", auth, timeout=15)
                if managers_data and "Members" in managers_data:
                    for manager in managers_data["Members"]:
                        manager_url = manager["@odata.id"]
                        manager_data = _rf_get(base_url, manager_url, auth, timeout=15)
                        if manager_data and "LogServices" in manager_data:
                            log_services_url = manager_data["LogServices"]["@odata.id"]
                            log_services_data = _rf_get(base_url, log_services_url, auth, timeout=15)
                            if log_services_data and "Members" in log_services_data:
                                available_logs = [service.get("Name", "Unknown") for service in log_services_data["Members"]]
                                return jsonify({
                                    "error": "LC logs not found. Available log services: " + ", ".join(available_logs),
                                    "available_logs": available_logs
                                }), 404
            except Exception:
                pass
            
            return jsonify({"error": "Could not fetch LC logs from iDRAC. Please check iDRAC version and permissions."}), 404

        # Extract log entries
        logs = []
        
        # Check if this is a direct log entries collection or a log service
        if "Members" in lc_log_data:
            members = lc_log_data.get("Members", [])
            # Get last 50 logs (or all if less than 50)
            for member in members[-50:]:
                try:
                    if "@odata.id" in member:
                        log_entry = _rf_get(base_url, member["@odata.id"], auth, timeout=15)
                        if log_entry:
                            logs.append({
                                "created": log_entry.get("Created", log_entry.get("Date", "N/A")),
                                "severity": log_entry.get("Severity", log_entry.get("EntryType", "INFO")),
                                "message": log_entry.get("Message", log_entry.get("Description", "N/A")),
                                "source": log_entry.get("Source", log_entry.get("OriginOfCondition", "N/A"))
                            })
                except Exception as e:
                    # Skip individual log entries that fail but continue processing others
                    continue
        elif "Entries" in lc_log_data:
            # Direct log entries collection
            entries = lc_log_data.get("Entries", [])
            for entry in entries[-50:]:
                logs.append({
                    "created": entry.get("Created", entry.get("Date", "N/A")),
                    "severity": entry.get("Severity", entry.get("EntryType", "INFO")),
                    "message": entry.get("Message", entry.get("Description", "N/A")),
                    "source": entry.get("Source", entry.get("OriginOfCondition", "N/A"))
                })
        else:
            # Try to extract logs directly if no Members or Entries
            if isinstance(lc_log_data, list):
                for log_item in lc_log_data[-50:]:
                    logs.append({
                        "created": log_item.get("Created", log_item.get("Date", "N/A")),
                        "severity": log_item.get("Severity", log_item.get("EntryType", "INFO")),
                        "message": log_item.get("Message", log_item.get("Description", "N/A")),
                        "source": log_item.get("Source", log_item.get("OriginOfCondition", "N/A"))
                    })

        # Reverse to show newest first
        logs.reverse()

        # Add debug information
        return jsonify({
            "logs": logs,
            "debug_info": {
                "endpoint_used": lc_log_path_used,
                "total_logs_found": len(logs),
                "response_structure": list(lc_log_data.keys()) if isinstance(lc_log_data, dict) else "non-dict"
            }
        })

    except Exception as e:
        return jsonify({"error": f"Failed to fetch LC logs: {str(e)}"}), 500


@app.route("/api/health", methods=["POST"])
def fetch_health_metrics():
    """Fetch server health metrics from iDRAC."""
    try:
        data = request.get_json()
        host = data.get("host")
        username = data.get("username")
        password = data.get("password")

        if not all([host, username, password]):
            return jsonify({"error": "Missing credentials"}), 400

        auth = (username, password)
        base_url = f"https://{host}"

        health_data = {
            "power": [],
            "temperature": [],
            "fan": [],
            "cpu": []
        }

        # Fetch Power information
        power_data = _rf_get(base_url, "/redfish/v1/Chassis/System.Embedded.1/Power", auth, timeout=30)
        if power_data:
            power_supplies = power_data.get("PowerSupplies", [])
            for ps in power_supplies:
                if ps.get("Status", {}).get("State") == "Enabled":
                    power_watts = ps.get("PowerConsumedWatts", 0)
                    health_data["power"].append(power_watts)

        # Fetch Thermal information (temperature sensors)
        thermal_data = _rf_get(base_url, "/redfish/v1/Chassis/System.Embedded.1/Thermal", auth, timeout=30)
        if thermal_data:
            temperatures = thermal_data.get("Temperatures", [])
            for temp in temperatures:
                if temp.get("Status", {}).get("State") == "Enabled":
                    reading_c = temp.get("ReadingCelsius", 0)
                    if reading_c > 0:
                        health_data["temperature"].append(reading_c)

        # Fetch Fan information
        if thermal_data:
            fans = thermal_data.get("Fans", [])
            for fan in fans:
                if fan.get("Status", {}).get("State") == "Enabled":
                    rpm = fan.get("Reading", 0)
                    if rpm > 0:
                        health_data["fan"].append(rpm)

        # Fetch Processor information (for CPU usage estimation)
        processors_data = _rf_get(base_url, "/redfish/v1/Systems/System.Embedded.1/Processors", auth, timeout=30)
        if processors_data:
            processors = processors_data.get("Members", [])
            for proc in processors[:4]:  # Limit to first 4 processors
                proc_data = _rf_get(base_url, proc["@odata.id"], auth, timeout=30)
                if proc_data:
                    # Estimate CPU usage from status (this is a simplified approach)
                    # Real CPU usage may require different endpoints
                    status = proc_data.get("Status", {}).get("Health", "Unknown")
                    if status == "OK":
                        health_data["cpu"].append(50)  # Placeholder for healthy CPU
                    else:
                        health_data["cpu"].append(0)

        # Calculate averages
        avg_power = sum(health_data["power"]) / len(health_data["power"]) if health_data["power"] else 0
        avg_temp = sum(health_data["temperature"]) / len(health_data["temperature"]) if health_data["temperature"] else 0
        avg_fan = sum(health_data["fan"]) / len(health_data["fan"]) if health_data["fan"] else 0
        avg_cpu = sum(health_data["cpu"]) / len(health_data["cpu"]) if health_data["cpu"] else 0

        # Get overall system health from iDRAC
        system_health = "Unknown"
        system_data = _rf_get(base_url, "/redfish/v1/Systems/System.Embedded.1", auth, timeout=30)
        if system_data:
            system_health = system_data.get("Status", {}).get("Health", "Unknown")

        # Get chassis health
        chassis_health = "Unknown"
        chassis_data = _rf_get(base_url, "/redfish/v1/Chassis/System.Embedded.1", auth, timeout=30)
        if chassis_data:
            chassis_health = chassis_data.get("Status", {}).get("Health", "Unknown")

        # Determine overall health status
        overall_health = "OK"
        if system_health in ["Critical", "Failed"] or chassis_health in ["Critical", "Failed"]:
            overall_health = "Critical"
        elif system_health in ["Warning", "Degraded"] or chassis_health in ["Warning", "Degraded"]:
            overall_health = "Warning"

        return jsonify({
            "metrics": {
                "power": round(avg_power, 1),
                "temperature": round(avg_temp, 1),
                "fan": round(avg_fan, 0),
                "cpu": round(avg_cpu, 1)
            },
            "health_status": overall_health,
            "system_health": system_health,
            "chassis_health": chassis_health,
            "raw": health_data
        })

    except Exception as e:
        return jsonify({"error": f"Failed to fetch health metrics: {str(e)}"}), 500


@app.route("/api/inventory-comparison", methods=["POST"])
def compare_inventory_interfaces():
    """Compare inventory data across Redfish, racadm, and IPMI interfaces."""
    try:
        data = request.get_json()
        host = data.get("host")
        username = data.get("username")
        password = data.get("password")

        if not all([host, username, password]):
            return jsonify({"error": "Missing credentials"}), 400

        # Fetch Redfish inventory (existing functionality)
        inventory_result, status, message = fetch_inventory(host, username, password)
        if status != "ok":
            return jsonify({"error": f"Failed to fetch Redfish inventory: {message}"}), 400

        redfish_inventory = inventory_result.get("inventory", [])
        normalized_redfish = _normalize_inventory_data(redfish_inventory, "redfish")

        # Fetch racadm hwinventory
        racadm_inventory = _parse_racadm_hwinventory(host, username, password)

        # Fetch IPMI FRU
        ipmi_inventory = _parse_ipmi_fru(host, username, password)

        # Compare across all three interfaces
        comparison_results = _compare_inventory_across_interfaces(
            normalized_redfish,
            racadm_inventory,
            ipmi_inventory
        )

        return jsonify({
            "comparison": comparison_results,
            "summary": {
                "redfish_count": len(normalized_redfish),
                "racadm_count": len(racadm_inventory),
                "ipmi_count": len(ipmi_inventory),
                "total_compared": len(comparison_results)
            }
        })

    except Exception as e:
        return jsonify({"error": f"Failed to compare inventory: {str(e)}"}), 500


@app.route("/api/fans", methods=["POST"])
def fetch_fan_details():
    """Fetch fan details from iDRAC web UI Hardware Inventory page."""
    try:
        data = request.get_json()
        host = data.get("host")
        username = data.get("username")
        password = data.get("password")

        if not all([host, username, password]):
            return jsonify({"error": "Missing credentials"}), 400

        import requests
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            return jsonify({"error": "beautifulsoup4 library not installed. Run: pip install beautifulsoup4"}), 500

        # Create session for web UI login
        session = requests.Session()
        session.verify = False  # Ignore SSL warnings for iDRAC
        requests.packages.urllib3.disable_warnings()

        # Login to iDRAC web UI
        login_url = f"https://{host}/index.html"
        login_data = {
            "user": username,
            "password": password
        }

        try:
            response = session.post(login_url, data=login_data, timeout=30, allow_redirects=True)
            if response.status_code != 200:
                return jsonify({"error": "Failed to login to iDRAC web UI"}), 401
        except Exception as e:
            return jsonify({"error": f"Login failed: {str(e)}"}), 401

        # Navigate to Hardware Inventory page
        # The URL pattern for hardware inventory is typically /sysinv.html or similar
        inventory_url = f"https://{host}/sysinv.html"
        try:
            response = session.get(inventory_url, timeout=30)
            if response.status_code != 200:
                # Try alternative URL patterns
                for alt_url in [f"https://{host}/cgi-bin/webcgi/sysinv", f"https://{host}/sysinv"]:
                    try:
                        response = session.get(alt_url, timeout=30)
                        if response.status_code == 200:
                            inventory_url = alt_url
                            break
                    except:
                        continue
                else:
                    return jsonify({"error": "Failed to access Hardware Inventory page"}), 404
        except Exception as e:
            return jsonify({"error": f"Failed to access inventory page: {str(e)}"}), 404

        # Parse HTML to extract fan descriptions
        soup = BeautifulSoup(response.text, 'html.parser')
        fans_data = []

        # Look for fan-related table rows or data
        # This depends on the specific HTML structure of iDRAC web UI
        # Common patterns: tables with "Fan" in headers, or specific data attributes

        # Try to find table with fan information
        tables = soup.find_all('table')
        for table in tables:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    text_content = ' '.join(cell.get_text(strip=True) for cell in cells)
                    # Look for rows containing fan information
                    if 'fan' in text_content.lower():
                        # Extract device description (second column typically)
                        if len(cells) >= 2:
                            description = cells[1].get_text(strip=True)
                            if description and description.lower() != 'fan':
                                fans_data.append({
                                    "description": description,
                                    "speed": "N/A",  # Speed might not be in inventory page
                                    "tier": "Unknown"
                                })

        # If no fans found in tables, try looking for specific data elements
        if not fans_data:
            # Look for divs or spans with fan data
            fan_elements = soup.find_all(text=lambda x: x and 'fan' in x.lower())
            for elem in fan_elements:
                parent = elem.parent
                if parent:
                    description = parent.get_text(strip=True)
                    if description and len(description) > 3 and description.lower() != 'fan':
                        if not any(f["description"] == description for f in fans_data):
                            fans_data.append({
                                "description": description,
                                "speed": "N/A",
                                "tier": "Unknown"
                            })

        # Determine fan tier from system model using Redfish (since web UI might not have it)
        try:
            system_data = _rf_get(f"https://{host}", "/redfish/v1/Systems/System.Embedded.1", (username, password), timeout=30)
            system_model = system_data.get("Model", "") if system_data else ""

            fan_tier = "Silver"
            if system_model:
                model_upper = system_model.upper()
                if any(x in model_upper for x in ["R750", "R760", "R7515", "R7615", "R840", "R850", "R960", "R960XA"]):
                    fan_tier = "Platinum"
                elif any(x in model_upper for x in ["R640", "R650", "R6515", "R6525", "R740", "R7415", "R7425"]):
                    fan_tier = "Gold"

            # Update tier for all fans
            for fan in fans_data:
                fan["tier"] = fan_tier
        except:
            pass  # If Redfish fails, keep Unknown tier

        return jsonify({"fans": fans_data, "system_model": system_model if 'system_model' in locals() else ""})

    except Exception as e:
        return jsonify({"error": f"Failed to fetch fan details: {str(e)}"}), 500


@app.route("/api/download-template")
def download_template():
    """Generate and serve a sample BOM Excel template matching 'Lab Build Sheet' format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Build Sheet"

    headers = ["Component Type", "ASSY DPN", "Part Number", "Quantity", "Slot", "Description"]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    # Sample rows
    samples = [
        ["CPU",  "VJWVJ",  "0VJWVJ",  2, "CPU.Socket.1", "Intel Xeon Platinum"],
        ["DIMM", "EXAMPLE-MEM-PN", "0EXAMPLE-MEM-PN", 16, "DIMM.Socket.A1", "64GB DDR5 4800MHz"],
        ["GPU",  "EXAMPLE-GPU-PN", "", 4, "Slot 3", "NVIDIA A100 80GB"],
        ["Disk", "EXAMPLE-SSD-PN", "", 8, "Disk.Bay.0", "960GB SAS SSD"],
        ["NIC",  "EXAMPLE-NIC-PN", "", 2, "", "Broadcom 25GbE"],
        ["PSU",  "EXAMPLE-PSU-PN", "", 2, "PSU.Slot.1", "2400W Power Supply"],
        ["PERC", "EXAMPLE-RAID-PN", "", 1, "", "PERC H755 Front"],
        ["Fan",  "EXAMPLE-FAN-PN", "", 6, "Fan.Bay.1", "Standard Cooling Fan"],
    ]
    for s in samples:
        ws.append(s)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bom_template.xlsx",
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Run on all network interfaces so team members can access via your IP
    app.run(host="0.0.0.0", port=5000, debug=True)
